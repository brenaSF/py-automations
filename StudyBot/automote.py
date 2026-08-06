# ==============================================================================
# SCRIPT DE AUTOMATIZAÇÃO DO CICLO DE ESTUDOS
# Gera/Atualiza a planilha excel do Ciclo de Estudos, Projeto Pessoal e Inglês
# ==============================================================================

import datetime
import os
import openpyxl
import requests
import dotenv
import telebot
from flask import Flask, request, jsonify
from services.sheetService import SheetsService

dotenv.load_dotenv()

app = Flask(__name__)

@app.route("/")
def home():
    return "Bot do Ciclo de Estudos está online!", 200


@app.route("/webhook", methods=["POST"])
def telegram_webhook():
    """Recebe as atualizações do Telegram e passa para o telebot processar."""
    if request.headers.get("content-type") == "application/json":
        json_string = request.get_data().decode("utf-8")
        update = telebot.types.Update.de_json(json_string)
        bot.process_new_updates([update])
        return "OK", 200
    return "Forbidden", 403


def configurar_webhook():
    """Registra a URL do Webhook junto ao Telegram na inicialização."""
    if WEBHOOK_URL:
        full_webhook_url = f"{WEBHOOK_URL.rstrip('/')}/webhook"
        bot.remove_webhook()
        sucesso = bot.set_webhook(url=full_webhook_url)
        if sucesso:
            print(f"[Info] Webhook configurado com sucesso: {full_webhook_url}")
        else:
            print("[Aviso] Falha ao configurar o Webhook no Telegram.")
    else:
        print("[Aviso] WEBHOOK_URL não configurada no arquivo .env")

DIR_BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_ARQUIVO= os.path.join(DIR_BASE , "ciclo_de_estudos_automatizado.xlsx")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")
WEBHOOK_URL = os.getenv("WEBHOOK_URL")

sheets_service = SheetsService()
bot = telebot.TeleBot(TELEGRAM_BOT_TOKEN)


def notificar_usuario_telegram(mensagem):
    """Envia uma notificação para o usuário via Telegram."""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": mensagem
    }
    
    try:
        response = requests.post(url, data=payload, timeout=30)

        response.raise_for_status()
        print("[Info] Notificação enviada via Telegram com sucesso!")
    except requests.exceptions.RequestException as e:
        print(f"[Aviso] Falha ao enviar mensagem pelo Telegram: {e}")

def registrar_sessao_estudo(data_str, categoria, materia, tempo_horas, questoes, acertos, observacao=""):
    """
    Adiciona um novo registro diretamente na planilha Excel usando openpyxl (sem dependência do Pandas).
    """
    if not os.path.exists(EXCEL_ARQUIVO):
        print(f"[Erro] Arquivo não encontrado em: {EXCEL_ARQUIVO}")
        print("Certifique-se de que o arquivo 'ciclo_de_estudos_automatizado.xlsx' está na mesma pasta.")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_ARQUIVO)
        
        pagina_planilha = "Diário de Estudos"
        if pagina_planilha not in wb.sheetnames:
            print(f"[Erro] Aba '{pagina_planilha}' não foi encontrada na planilha.")
            return

        ws = wb[pagina_planilha]

        taxa = (acertos / questoes) if questoes > 0 else 0.0

        nova_linha = [data_str, categoria, materia, tempo_horas, questoes, acertos, taxa, observacao]
        ws.append(nova_linha)

        last_row = ws.max_row
        ws.cell(row=last_row, column=4).number_format = '0.0 "h"'
        ws.cell(row=last_row, column=5).number_format = '#,##0'
        ws.cell(row=last_row, column=6).number_format = '#,##0'
        ws.cell(row=last_row, column=7).number_format = '0.0%'

        wb.save(EXCEL_ARQUIVO)
        print(f"\n[Sucesso] Registro salvo na planilha!")
        print(f"Adicionado: {data_str} | {categoria} | {materia} | {tempo_horas}h | {acertos}/{questoes}")

        return True,nova_linha

    except PermissionError:
        print("\n[Atenção] O arquivo Excel está aberto! Feche a planilha e execute o script novamente.")

        return False
    except Exception as e:
        print(f"\n[Erro ao atualizar a planilha]: {e}")
        return False

def registrar_sessao_estudo_online(data_str, categoria, materia, tempo_horas, questoes, acertos, taxa, observacao):
    sucesso = sheets_service.adicionar_sessao_estudo(
        data_str, categoria, materia, tempo_horas, questoes, acertos, taxa, observacao
    )
    return sucesso, [data_str, categoria, materia, tempo_horas, questoes, acertos, observacao]

def formatar_resposta(lista):

    data_str = lista[0]
    categoria = lista[1]
    materia = lista[2]
    tempo_horas = lista[3]
    questoes = lista[4]
    acertos = lista[5]
    observacao = lista[6]

    try:
        tempo_horas_num = float(tempo_horas)
        tempo_minutos = int(tempo_horas_num * 60)
    except (ValueError, TypeError):
        tempo_horas_num = 0.0
        tempo_minutos = 0

    taxa = (acertos / questoes) if questoes > 0 else 0.0
    taxa_pct = taxa * 100

    if taxa_pct >= 80:
        status = "🔥 *EXCELENTE!* (Acima da média)"
    elif taxa_pct >= 60:
        status = "📈 *BOM DESEMPENHO!*"
    else:
        status = "⚠️ *ATENÇÃO!* (Precisa revisar)"

    blocos_verdes = int(taxa_pct // 10)
    barra_progresso = "🟩" * blocos_verdes + "⬜" * (10 - blocos_verdes)

    return f"🏆 *SESSÃO REGISTRADA*\n\n" \
           f"🎮 *Matéria:* {materia} ({categoria})\n" \
           f"⏱️ *Investido:* {tempo_minutos} min ({tempo_horas}h)\n\n" \
           f"📈 *Resultado:* **{taxa_pct:.1f}% de Aproveitamento** ({acertos}/{questoes})\n" \
           f"Status: {status}\n" \
           f"{barra_progresso}\n\n" \
           f"💬 *Nota:*\n" \
           f"_{observacao}_\n\n" \
           f"⏭️ *Próximo Alvo:* Dar continuidade ao ciclo!"

dados_usuario = {}

@bot.message_handler(commands=['estudo'])
def iniciar_registro(message):
    user_id = message.chat.id
    dados_usuario[user_id] = {}  # Limpa/inicializa os dados do usuário
    
    msg = bot.send_message(user_id, "📚 *NOVO REGISTRO DE ESTUDO*\n\nQual é a *categoria*? (ex: Concurso, Faculdade, Inglês)", parse_mode="Markdown")
    bot.register_next_step_handler(msg, obter_categoria)

def obter_categoria(message):
    user_id = message.chat.id
    dados_usuario[user_id]['categoria'] = message.text
    
    msg = bot.send_message(user_id, "📖 Qual é a *matéria/tópico* estudado? (ex: Java - Streams API)", parse_mode="Markdown")
    bot.register_next_step_handler(msg, obter_materia)

def obter_materia(message):
    user_id = message.chat.id
    dados_usuario[user_id]['materia'] = message.text
    
    msg = bot.send_message(user_id, "⏱️ Quantas *horas* você estudou? (ex: 2 ou 1.5)", parse_mode="Markdown")
    bot.register_next_step_handler(msg, obter_tempo)

def obter_tempo(message):
    user_id = message.chat.id
    try:
        dados_usuario[user_id]['tempo_horas'] = float(message.text.replace(',', '.'))
    except ValueError:
        msg = bot.send_message(user_id, "⚠️ Valor inválido! Digite apenas números para o tempo em horas (ex: 2 ou 1.5):")
        bot.register_next_step_handler(msg, obter_tempo)
        return

    msg = bot.send_message(user_id, "❓ Quantas *questões* você resolveu? (digite 0 se não fez questões)", parse_mode="Markdown")
    bot.register_next_step_handler(msg, obter_questoes)

def obter_questoes(message):
    user_id = message.chat.id
    try:
        dados_usuario[user_id]['questoes'] = int(message.text)
    except ValueError:
        msg = bot.send_message(user_id, "⚠️ Valor inválido! Digite um número inteiro para as questões:")
        bot.register_next_step_handler(msg, obter_questoes)
        return

    if dados_usuario[user_id]['questoes'] > 0:
        msg = bot.send_message(user_id, "🎯 Quantas questões você *acertou*?", parse_mode="Markdown")
        bot.register_next_step_handler(msg, obter_acertos)
    else:
        dados_usuario[user_id]['acertos'] = 0
        msg = bot.send_message(user_id, "💬 Alguma *observação/anotação* sobre a sessão? (ou envie '-' se não houver)", parse_mode="Markdown")
        bot.register_next_step_handler(msg, obter_observacao)

def obter_acertos(message):
    user_id = message.chat.id
    try:
        acertos = int(message.text)
        questoes = dados_usuario[user_id]['questoes']
        if acertos > questoes:
            msg = bot.send_message(user_id, f"⚠️ O número de acertos ({acertos}) não pode ser maior que o total de questões ({questoes}). Tente novamente:")
            bot.register_next_step_handler(msg, obter_acertos)
            return
        dados_usuario[user_id]['acertos'] = acertos
    except ValueError:
        msg = bot.send_message(user_id, "⚠️ Valor inválido! Digite um número inteiro para os acertos:")
        bot.register_next_step_handler(msg, obter_acertos)
        return

    msg = bot.send_message(user_id, "💬 Alguma *observação/anotação* sobre a sessão? (ou envie '-' se não houver)", parse_mode="Markdown")
    bot.register_next_step_handler(msg, obter_observacao)

def obter_observacao(message):
    user_id = message.chat.id
    dados_usuario[user_id]['observacao'] = message.text if message.text != '-' else ""
    
    dados = dados_usuario[user_id]
    hoje = datetime.date.today().strftime("%Y-%m-%d")

    bot.send_message(user_id, "⏳ Registrando sua sessão na planilha...")

    sucesso, lista = registrar_sessao_estudo_online(
        data_str=hoje,
        categoria=dados['categoria'],
        materia=dados['materia'],
        tempo_horas=dados['tempo_horas'],
        questoes=dados['questoes'],
        acertos=dados['acertos'],
        taxa=(dados['acertos'] / dados['questoes']) if dados['questoes'] > 0 else 0.0,
        observacao=dados['observacao']
    )

    if sucesso:
        bot.send_message(user_id, formatar_resposta(lista), parse_mode="Markdown")
    else:
        bot.send_message(user_id, '❌ Ocorreu um erro ao registrar a sessão de estudo na planilha.')
    
    del dados_usuario[user_id]

ENV = os.getenv("ENV", "local")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))

    if ENV == "production" and WEBHOOK_URL:
        configurar_webhook()
        app.run(host="0.0.0.0", port=port)
    else:
        print("[Info] Rodando em modo LOCAL via Polling...")
        bot.remove_webhook()
        
        import threading
        bot_thread = threading.Thread(target=bot.infinity_polling, daemon=True)
        bot_thread.start()

        app.run(host="127.0.0.1", port=port, debug=False)